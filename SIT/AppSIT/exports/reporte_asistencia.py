from io import BytesIO
from openpyxl import load_workbook
import os
from AppSIT.models import TutoringSession, TutoringAttendance, Group, StudentGroup, User

def mapear_hojas(wb):
    """
    Mapea las hojas 'Individual' y 'Grupal' de forma dinámica.
    """
    hojas = {
        "Individual": next((sheet for sheet in wb.sheetnames if "individual" in sheet.lower()), None),
        "Grupal": next((sheet for sheet in wb.sheetnames if "grupal" in sheet.lower()), None),
    }
    if not hojas["Individual"]:
        raise ValueError("La hoja 'Individual' no existe en el archivo.")
    if not hojas["Grupal"]:
        raise ValueError("La hoja 'Grupal' no existe en el archivo.")
    return hojas


def set_value(worksheet, cell_address, value):
    """
    Asigna valores a celdas, manejando combinaciones.
    """
    if value is None:
        return  # No asignar valores nulos
    cell = worksheet[cell_address]
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left_cell = merged_range.coord.split(":")[0]
            worksheet[top_left_cell].value = value
            return
    cell.value = value


def rellenar_datos_generales(hoja, tutor, grupo):
    """
    Rellena los datos generales comunes en una hoja.
    """
    academias = {
        "ISC": "Ingeniería en Sistemas Computacionales",
        "IM": "Ingeniería Mecatrónica",
        "IB": "Ingeniería Bioquímica",
        "II": "Ingeniería Industrial",
        "LG": "Licenciatura en Gastronomía",
        "N/A": "No aplica",
    }
    total_alumnos = StudentGroup.objects.filter(group=grupo).count()
    mujeres = User.objects.filter(genero="Femenino", studentgroup__group=grupo).count()
    hombres = User.objects.filter(genero="Masculino", studentgroup__group=grupo).count()

    set_value(hoja, "F2", tutor.nombrecompleto())
    set_value(hoja, "E3", f"{grupo.semester} {grupo.group_name}")
    set_value(hoja, "D4", total_alumnos)
    set_value(hoja, "D5", mujeres)
    set_value(hoja, "I5", hombres)


def rellenar_porcentajes_asistencia_individual(hoja, sesiones):
    """
    Rellena los porcentajes de asistencia en la hoja Individual.
    """
    for idx, sesion in enumerate(sesiones, start=1):
        col_fecha = 4 + (idx - 1) * 2  # D=4, F=6, H=8...
        if col_fecha > 34:  # Máximo columna AH
            break

        # Fecha de la sesión
        fecha_sesion = sesion.date.strftime("%d/%m/%Y") if sesion.date else ""
        set_value(hoja, f"{chr(64 + col_fecha)}10", fecha_sesion)

        # Total de estudiantes presentes
        total_presentes = sesion.attendances.filter(is_present=True).count()
        set_value(hoja, f"{chr(64 + col_fecha)}11", total_presentes)

        # Conteo de hombres y mujeres en la sesión
        mujeres_sesion = sesion.attendances.filter(is_present=True, student__genero="Femenino").count()
        hombres_sesion = sesion.attendances.filter(is_present=True, student__genero="Masculino").count()

        # Rellenar en las columnas correctas
        set_value(hoja, f"{chr(64 + col_fecha)}13", mujeres_sesion)  # Mujeres (D)
        set_value(hoja, f"{chr(64 + col_fecha + 1)}13", hombres_sesion)  # Hombres (E)


def rellenar_porcentajes_asistencia_grupal(hoja, sesiones):
    """
    Rellena los porcentajes de asistencia en la hoja Grupal.

    Parámetros:
    - hoja: La hoja de Excel a rellenar.
    - sesiones: QuerySet de sesiones grupales.
    """
    max_col = 33  # AG es la 33 en índice de Excel (1-based)
    fila_fecha = 10
    fila_hombres = 12
    fila_mujeres = 12

    for idx, sesion in enumerate(sesiones, start=1):
        col_inicio = 3 + (idx - 1) * 2  # C=3, E=5, G=7...

        if col_inicio > max_col:
            break  # Salimos si excede AG

        # Fecha de la sesión
        set_value(hoja, f"{chr(64 + col_inicio)}{fila_fecha}", sesion.date.strftime("%d/%m/%Y"))

        # Conteo de hombres y mujeres presentes
        hombres_sesion = sesion.attendances.filter(is_present=True, student__genero="Masculino").count()
        mujeres_sesion = sesion.attendances.filter(is_present=True, student__genero="Femenino").count()

        # Rellenar número de hombres y mujeres
        set_value(hoja, f"{chr(64 + col_inicio + 1)}{fila_hombres}", hombres_sesion)  # D12: Hombres
        set_value(hoja, f"{chr(64 + col_inicio)}{fila_mujeres}", mujeres_sesion)  # C12: Mujeres



def rellenar_nombres_estudiantes(hoja, sesiones, fila_inicio, fila_fin):
    """
    Rellena los nombres de los estudiantes en las filas asignadas.
    """
    fila_actual = fila_inicio
    for sesion in sesiones:
        for asistencia in sesion.attendances.filter(is_present=True):
            if fila_actual > fila_fin:
                break
            set_value(hoja, f"D{fila_actual}", asistencia.student.nombrecompleto())
            fila_actual += 1


def generar_reporte_asistencia(tutor):
    """
    Genera el reporte de asistencia individual y grupal para un tutor.
    """
    try:
        grupo = Group.objects.get(teacher=tutor)
    except Group.DoesNotExist:
        raise ValueError("No tienes un grupo asignado.")

    # Construir la ruta al archivo base
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path_input_excel = os.path.join(base_dir, "exports", "R14-PC18 Reporte de asistencia a tutoría Individual - Grupal.xlsx")

    if not os.path.exists(path_input_excel):
        raise FileNotFoundError(f"No se encontró el archivo: {path_input_excel}")

    # Cargar el archivo base y mapear las hojas
    wb = load_workbook(path_input_excel)
    hojas = mapear_hojas(wb)
    hoja_individual = wb[hojas["Individual"]]
    hoja_grupal = wb[hojas["Grupal"]]

    # --- Hoja Individual ---
    rellenar_datos_generales(hoja_individual, tutor, grupo)
    sesiones_individuales = TutoringSession.objects.filter(
        tutor=tutor, is_group=False
    ).prefetch_related("attendances")
    rellenar_porcentajes_asistencia_individual(hoja_individual, sesiones_individuales)
    rellenar_nombres_estudiantes(hoja_individual, sesiones_individuales, 19, 23)

    # --- Hoja Grupal ---
    rellenar_datos_generales(hoja_grupal, tutor, grupo)
    sesiones_grupales = TutoringSession.objects.filter(
        tutor=tutor, is_group=True
    ).prefetch_related("attendances")
    rellenar_porcentajes_asistencia_grupal(hoja_grupal, sesiones_grupales)

    # Guardar en un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
