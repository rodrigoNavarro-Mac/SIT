import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

def generar_reporte_seguimiento():
    # Ruta del archivo Excel
    base_dir = os.path.dirname(os.path.abspath(__file__))  # Directorio del archivo actual
    input_file = os.path.join(base_dir, 'R04-PC18 Seguimiento académico-1A-ISC.xlsx')

    # Validar existencia del archivo
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"El archivo no existe: {input_file}")

    # Cargar el archivo Excel
    wb_input = openpyxl.load_workbook(input_file)

    # Buscar la hoja "1er Parcial" ignorando mayúsculas/minúsculas
    hoja_esperada = "1er Parcial"
    ws_input = None
    for sheet_name in wb_input.sheetnames:
        if hoja_esperada.lower() == sheet_name.lower():  # Comparación insensible a mayúsculas/minúsculas
            ws_input = wb_input[sheet_name]
            break

    if ws_input is None:
        raise ValueError(f"La hoja '{hoja_esperada}' no existe en el archivo. Hojas disponibles: {wb_input.sheetnames}")

    # Crear un nuevo libro de trabajo para el reporte
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Reporte Seguimiento"

    # Estilos
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezado principal
    ws_output.merge_cells("A1:E1")
    ws_output["A1"].value = "Reporte de Seguimiento - 1er Parcial"
    ws_output["A1"].font = bold_font
    ws_output["A1"].alignment = center_alignment

    # Copiar encabezados desde la hoja original
    headers = [cell.value for cell in ws_input[1]]
    ws_output.append(headers)

    # Aplicar estilos a los encabezados
    for col_num, header in enumerate(headers, start=1):
        cell = ws_output.cell(row=2, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Copiar datos desde la hoja original
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        ws_output.append(row)

    # Aplicar estilos a las celdas del reporte
    for row_num in range(3, ws_output.max_row + 1):
        for col_num in range(1, ws_output.max_column + 1):
            cell = ws_output.cell(row=row_num, column=col_num)
            cell.border = thin_border

    # Ajuste de anchura de columnas para mejor visibilidad
    for col in ws_output.columns:
        # Ignorar celdas fusionadas
        max_length = max(
            len(str(cell.value or "")) for cell in col if not isinstance(cell, openpyxl.cell.cell.MergedCell)
        )
        col_letter = openpyxl.utils.get_column_letter(col[0].column)  # Convertir número de columna a letra
        ws_output.column_dimensions[col_letter].width = max_length + 2

    return wb_output
