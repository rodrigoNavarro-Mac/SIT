from io import BytesIO
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font
from AppSIT.models import Grade, Group, StudentGroup, StudentSubject, TeacherSubject
import os


def obtener_grupo_tutor(tutor):
    """
    Obtiene el grupo asignado al tutor.
    """
    try:
        grupo = Group.objects.get(teacher=tutor)
        return grupo
    except Group.DoesNotExist:
        return None


def obtener_alumnos_grupo(grupo):
    """
    Obtiene los alumnos asignados a un grupo.
    """
    return StudentGroup.objects.filter(group=grupo).select_related('student')


def obtener_materias_alumno(alumno):
    """
    Obtiene las materias que cursa un alumno.
    """
    return StudentSubject.objects.filter(student=alumno).select_related('subject')


def obtener_calificaciones_materia(alumno, materia):
    """
    Obtiene las calificaciones parciales de un alumno en una materia.
    """
    try:
        return Grade.objects.get(
            student_subject__student=alumno,
            student_subject__subject=materia
        )
    except Grade.DoesNotExist:
        return None

def rellenar_informacion_general(hoja, tutor, grupo, alumnos, hoja_nombre):
    """
    Rellena la tabla de Información General en la hoja actual.
    """
    # Mapear códigos de academias a nombres completos
    academias = {
        "ISC": "Ingeniería en Sistemas Computacionales",
        "IM": "Ingeniería Mecatrónica",
        "IB": "Ingeniería Bioquímica",
        "II": "Ingeniería Industrial",
        "LG": "Licenciatura en Gastronomía",
        "N/A": "No aplica",
    }

    # Obtener el nombre completo de la academia
    carrera = academias.get(tutor.academia, "Desconocida")

    # Rellenar las celdas específicas
    hoja["P3"] = carrera  # Carrera
    hoja["P4"] = tutor.nombrecompleto()  # Persona tutora
    hoja["P5"] = hoja_nombre  # Parcial
    hoja["P6"] = f"{grupo.semester} {grupo.group_name}"  # Semestre y grupo
    hoja["R7"] = len(alumnos)  # No. total de estudiantes

    # Depuración
    print("Información General Rellenada:")
    print(f"Carrera: {carrera}")
    print(f"Persona tutora: {tutor.nombrecompleto()}")
    print(f"Parcial: {hoja_nombre}")
    print(f"Semestre y grupo: {grupo.semester} {grupo.group_name}")
    print(f"No. total de estudiantes: {len(alumnos)}")

def rellenar_informacion_academica(hoja, grupo, alumnos):
    """
    Rellena la tabla de Información Académica en la hoja actual.

    Parámetros:
    - hoja: La hoja actual del archivo Excel.
    - grupo: El objeto Group que representa el grupo del tutor.
    - alumnos: La lista de alumnos del grupo.
    """
    # Obtener las materias únicas del grupo
    materias = set()
    materias_docentes = {}  # Relación de materia y docente
    materias_estudiantes = {}  # Conteo de estudiantes por materia

    # Construir las materias, docentes y conteo de estudiantes
    for alumno_grupo in alumnos:
        alumno = alumno_grupo.student
        materias_asignadas = StudentSubject.objects.filter(student=alumno)

        for materia_asignada in materias_asignadas:
            materia = materia_asignada.subject
            docente_materia = TeacherSubject.objects.filter(subject=materia).first()
            if not docente_materia:
                continue

            # Agregar materia y docente
            materias.add(materia.subject_name)
            materias_docentes[materia.subject_name] = docente_materia.teacher.nombrecompleto()
            materias_estudiantes[materia.subject_name] = materias_estudiantes.get(materia.subject_name, 0) + 1

    # Ordenar materias por nombre
    materias_ordenadas = sorted(materias)

    # Rellenar las celdas de la tabla
    for idx, materia in enumerate(materias_ordenadas):
        # Verificar rango (de AB4:AL4 hasta AB23:AL23)
        fila_inicio = 4 + idx
        if fila_inicio > 23:
            break  # Evitar sobrescribir fuera del rango permitido

        # Rellenar Nombre de la Asignatura
        celda_asignatura = f"AB{fila_inicio}"
        hoja[celda_asignatura].value = materia

        # Rellenar Nombre del Docente
        celda_docente = f"AM{fila_inicio}"
        hoja[celda_docente].value = materias_docentes.get(materia, "Desconocido")

        # Rellenar Número Total de Estudiantes
        celda_estudiantes = f"AW{fila_inicio}"
        hoja[celda_estudiantes].value = materias_estudiantes.get(materia, 0)

    # Depuración
    print("Información Académica Rellenada:")
    for idx, materia in enumerate(materias_ordenadas):
        print(f"{idx + 1}. Asignatura: {materia}, Docente: {materias_docentes.get(materia)}, Estudiantes: {materias_estudiantes.get(materia)}")
       
       

def obtener_calificaciones_por_parcial(alumnos, materias, parcial_numero):
    """
    Obtiene las calificaciones por parcial para los alumnos y materias.

    Parámetros:
    - alumnos: Lista de estudiantes del grupo.
    - materias: Lista de materias únicas que toman los estudiantes.
    - parcial_numero: Número del parcial (1, 2, 3, 4).

    Retorna:
    - Un diccionario con las calificaciones por alumno y materia.
    """
    calificaciones = {}
    
    for alumno_grupo in alumnos:
        alumno = alumno_grupo.student
        calificaciones[alumno.username] = {}
        
        for materia in materias:
            grade = obtener_calificaciones_materia(alumno, materia)
            if grade:
                calificacion = getattr(grade, f"parcial_{parcial_numero}", 0)  # Obtener calificación del parcial
            else:
                calificacion = 0  # Si no hay calificación, se asigna 0
            calificaciones[alumno.username][materia.subject_name] = calificacion
    
    return calificaciones

 
        
def rellenar_tabla_detalle(hoja, alumnos, materias, calificaciones_por_parcial):
    """
    Rellena la tabla de detalle de estudiantes y calificaciones.

    Parámetros:
    - hoja: La hoja actual del archivo Excel.
    - alumnos: Lista de estudiantes en el grupo.
    - materias: Lista de materias únicas que toman los estudiantes.
    - calificaciones_por_parcial: Diccionario con las calificaciones por alumno y materia.
      Ejemplo: {alumno_username: {materia: calificacion}}
    """
    from openpyxl.utils import get_column_letter

    def validar_valor_excel(valor):
        """
        Valida que el valor sea compatible con Excel.
        Retorna el valor o lo convierte a cadena si no es compatible.
        """
        if isinstance(valor, (str, int, float, type(None))):
            return valor
        return str(valor)  # Convertir cualquier otro tipo a cadena

    # Rellenar No. de Matrícula y Nombre del Estudiante
    for fila, alumno_grupo in enumerate(alumnos, start=28):
        alumno = alumno_grupo.student
        # No. de Matrícula
        hoja[f"A{fila}"].value = validar_valor_excel(alumno.username)
        # Nombre del Estudiante
        hoja[f"B{fila}"].value = validar_valor_excel(alumno.nombrecompleto())
        # Género: Hombres o Mujeres
        if alumno.genero == "Masculino":
            hoja[f"O{fila}"].value = 1
        elif alumno.genero == "Femenino":
            hoja[f"P{fila}"].value = 1

    # Rellenar Materias y Calificaciones
    columna_inicial = 26  # Columna Z es la 26 en 1-indexed Excel
    for idx_materia, materia in enumerate(materias):
        # Nombre de la materia (Fila 25)
        columna_materia = columna_inicial + (idx_materia * 3)
        celda_materia = f"{get_column_letter(columna_materia)}25"
        hoja[celda_materia].value = validar_valor_excel(materia.subject_name)

        # Rellenar calificaciones por materia
        for fila, alumno_grupo in enumerate(alumnos, start=28):
            alumno = alumno_grupo.student
            calificacion = calificaciones_por_parcial.get(alumno.username, {}).get(materia.subject_name, 0)

            # Celdas aprobatorias y reprobatorias
            celda_aprobado = f"{get_column_letter(columna_materia)}{fila}"
            celda_reprobado = f"{get_column_letter(columna_materia + 1)}{fila}"

            if calificacion >= 70:
                hoja[celda_aprobado].value = 1  # Aprobado
                hoja[celda_reprobado].value = None  # No necesita marca en Reprobado
            else:
                hoja[celda_reprobado].value = 1  # Reprobado
                hoja[celda_aprobado].value = None  # No necesita marca en Aprobado

    # Depuración
    print("Tabla de Detalle Rellenada:")
    for fila, alumno_grupo in enumerate(alumnos, start=28):
        alumno = alumno_grupo.student
        print(f"Alumno: {alumno.username}, Género: {alumno.genero}")
        for idx_materia, materia in enumerate(materias):
            calificacion = calificaciones_por_parcial.get(alumno.username, {}).get(materia.subject_name, 0)
            print(f"  Materia: {materia.subject_name}, Calificación: {calificacion}")




def generar_reporte_seguimiento(tutor):
    """
    Genera un reporte en base al archivo Excel subido y lo llena con los datos del tutor.
    """
    # Verificar si el tutor tiene un grupo asignado
    grupo = obtener_grupo_tutor(tutor)
    if not grupo:
        raise ValueError("No tienes un grupo asignado.")

    # Obtener los alumnos del grupo
    alumnos = obtener_alumnos_grupo(grupo)

    # Construir la ruta absoluta al archivo base
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Directorio base del proyecto
    path_input_excel = os.path.join(base_dir, "exports", "R04-PC18 Seguimiento académico.xlsx")

    # Verificar que el archivo exista
    if not os.path.exists(path_input_excel):
        raise FileNotFoundError(f"No se encontró el archivo: {path_input_excel}")

    # Cargar el archivo base
    wb = load_workbook(path_input_excel)
    print(f"Hojas disponibles en el archivo: {wb.sheetnames}")

    # Mapeo de las hojas por parcial
    hojas_parciales = {
        1: next((sheet for sheet in wb.sheetnames if "1er parcial" in sheet.lower()), None),
        2: next((sheet for sheet in wb.sheetnames if "2do parcial" in sheet.lower()), None),
        3: next((sheet for sheet in wb.sheetnames if "3er parcial" in sheet.lower()), None),
        4: next((sheet for sheet in wb.sheetnames if "4to parcial" in sheet.lower()), None),
    }

    # Obtener las materias únicas del grupo
    materias = set()
    for alumno_grupo in alumnos:
        materias_asignadas = obtener_materias_alumno(alumno_grupo.student)
        for materia_asignada in materias_asignadas:
            materias.add(materia_asignada.subject)
    materias = sorted(materias, key=lambda m: m.subject_name)  # Ordenar materias por nombre

    # Rellenar cada hoja con los datos del parcial correspondiente
    for parcial_numero, hoja_nombre in hojas_parciales.items():
        if hoja_nombre not in wb.sheetnames:
            raise ValueError(f"La hoja '{hoja_nombre}' no existe en el archivo Excel.")

        hoja = wb[hoja_nombre]

        # Rellenar información general
        rellenar_informacion_general(hoja, tutor, grupo, alumnos, hoja_nombre)

        # Rellenar información académica
        rellenar_informacion_academica(hoja, grupo, alumnos)

        # Obtener calificaciones por parcial
        calificaciones_por_parcial = obtener_calificaciones_por_parcial(alumnos, materias, parcial_numero)

        # Rellenar tabla de detalle
        rellenar_tabla_detalle(hoja, alumnos, materias, calificaciones_por_parcial)

    # Guardar en un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
